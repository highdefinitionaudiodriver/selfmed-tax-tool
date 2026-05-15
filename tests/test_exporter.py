"""exporter.py のユニットテスト。"""

import sys
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from core.exporter import export_xlsx


@pytest.fixture
def sample_df():
    # quantity が 2 の行を含めて、支払金額が unit_price * quantity になることを検証
    unit_prices = [698, 1280, 1980]
    quantities = [1, 2, 1]
    paid_amounts = [u * q for u, q in zip(unit_prices, quantities)]
    return pd.DataFrame({
        "seller": ["Amazon", "Amazon", "ヘルスケアストア"],
        "product_name": ["ロキソニンS 12錠", "パブロンゴールドA キッズ", "アレグラFX 28錠"],
        "unit_price": unit_prices,
        "quantity": quantities,
        "paid_amount": paid_amounts,
        "order_date": pd.to_datetime(["2025-03-15", "2025-06-02", "2025-09-01"]),
        "判定": ["対象", "要確認", "対象"],
    })


@pytest.fixture
def output_path(tmp_path):
    return tmp_path / "test_output.xlsx"


class TestExportXlsx:
    def test_creates_file(self, sample_df, output_path):
        result = export_xlsx(sample_df, output_path)
        assert result.exists()

    def test_sheet_name(self, sample_df, output_path):
        export_xlsx(sample_df, output_path)
        wb = load_workbook(output_path)
        assert wb.sheetnames == ["セルフメディケーション明細"]

    def test_header_row(self, sample_df, output_path):
        export_xlsx(sample_df, output_path)
        wb = load_workbook(output_path)
        ws = wb.active
        headers = [ws.cell(row=1, column=c).value for c in range(1, 8)]
        assert headers == [
            "支払先の名称", "医薬品の名称", "単価", "数量",
            "支払った金額", "購入日", "判定",
        ]

    def test_data_rows(self, sample_df, output_path):
        export_xlsx(sample_df, output_path)
        wb = load_workbook(output_path)
        ws = wb.active
        # 3行のデータ（列2 = 医薬品の名称）
        assert ws.cell(row=2, column=2).value == "ロキソニンS 12錠"
        assert ws.cell(row=3, column=2).value == "パブロンゴールドA キッズ"
        assert ws.cell(row=4, column=2).value == "アレグラFX 28錠"
        # 数量 (列4) と 支払った金額 (列5) の整合性
        assert ws.cell(row=3, column=4).value == 2
        assert ws.cell(row=3, column=5).value == 1280 * 2

    def test_total_row_uses_paid_amount(self, sample_df, output_path):
        export_xlsx(sample_df, output_path)
        wb = load_workbook(output_path)
        ws = wb.active
        # 合計行: row=5 (ヘッダ1 + データ3 + 合計1)
        # 「合計」ラベルは paid_amount 列の1つ左（=数量列=列4）に置かれる
        assert ws.cell(row=5, column=4).value == "合計"
        # 合計は unit_price * quantity の合計（数量2の行が正しく反映されている）
        expected_total = 698 * 1 + 1280 * 2 + 1980 * 1
        assert ws.cell(row=5, column=5).value == expected_total

    def test_review_row_highlighted(self, sample_df, output_path):
        export_xlsx(sample_df, output_path)
        wb = load_workbook(output_path)
        ws = wb.active
        # 「要確認」行 (row=3) が黄色ハイライト
        fill_color = ws.cell(row=3, column=1).fill.start_color.rgb
        assert fill_color == "00FFF2CC"

    def test_empty_dataframe(self, output_path):
        empty_df = pd.DataFrame(columns=[
            "seller", "product_name", "unit_price", "quantity", "paid_amount", "order_date", "判定"
        ])
        export_xlsx(empty_df, output_path)
        wb = load_workbook(output_path)
        ws = wb.active
        # ヘッダ + 合計行のみ
        assert ws.cell(row=2, column=4).value == "合計"
        assert ws.cell(row=2, column=5).value == 0
