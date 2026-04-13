"""設計書（design_document.xlsx）生成スクリプト。

現在のソースコードの実装状態を正確に反映した設計書を生成する。
シート構成:
  1. 表紙
  2. 機能一覧表
  3. API仕様書（CLI + 内部モジュール関数）
  4. テーブル定義書（JSON設定ファイルのスキーマ）
  5. エラー・ログ定義書
  6. アーキテクチャ図解（Mermaid記法）
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# --- 共通スタイル ---
FONT_NAME = "Arial"
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
SECTION_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
ALT_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
THIN = Side(style="thin", color="808080")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

HEADER_FONT = Font(name=FONT_NAME, bold=True, size=11, color="FFFFFF")
BODY_FONT = Font(name=FONT_NAME, size=10)
SECTION_FONT = Font(name=FONT_NAME, bold=True, size=11)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=16)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)


def style_header_row(ws, row: int, num_cols: int) -> None:
    for col in range(1, num_cols + 1):
        c = ws.cell(row=row, column=col)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER


def style_body_row(ws, row: int, num_cols: int, alt: bool = False) -> None:
    for col in range(1, num_cols + 1):
        c = ws.cell(row=row, column=col)
        c.font = BODY_FONT
        c.alignment = LEFT
        c.border = BORDER
        if alt:
            c.fill = ALT_FILL


def set_column_widths(ws, widths: list[int]) -> None:
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def write_table(ws, start_row: int, headers: list[str], rows: list[list]) -> int:
    """ヘッダ + データ行を書き込み、次に書き込むべき行番号を返す。"""
    for col, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=col, value=h)
    style_header_row(ws, start_row, len(headers))

    for r_idx, row_data in enumerate(rows):
        r = start_row + 1 + r_idx
        for c_idx, value in enumerate(row_data, 1):
            ws.cell(row=r, column=c_idx, value=value)
        style_body_row(ws, r, len(headers), alt=(r_idx % 2 == 1))
    return start_row + 1 + len(rows)


# ============================================================
# シート 1: 表紙
# ============================================================
def build_cover(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "表紙"
    set_column_widths(ws, [20, 60])

    ws.merge_cells("A1:B3")
    c = ws.cell(row=1, column=1, value="Selfmed Tax Tool\n設計書")
    c.font = Font(name=FONT_NAME, bold=True, size=20, color="FFFFFF")
    c.alignment = CENTER
    c.fill = HEADER_FILL
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 30
    ws.row_dimensions[3].height = 30

    meta = [
        ("プロジェクト名", "selfmed-tax-tool"),
        ("プロジェクト概要", "ECサイト購入履歴CSVからセルフメディケーション税制対象品を抽出しxlsx出力するCLIツール"),
        ("バージョン", "1.0.0"),
        ("対象読者", "開発者 / 保守担当 / 外部ベンダー"),
        ("実装言語", "Python 3.10+"),
        ("主要依存", "pandas, openpyxl"),
        ("実行形態", "CLI（argparse） / PyInstaller EXE"),
        ("設計書作成日", "2026-04-11"),
        ("ソースコード基準", "main.py / core/loader.py / core/matcher.py / core/exporter.py"),
    ]
    start = 5
    for i, (k, v) in enumerate(meta):
        r = start + i
        ws.cell(row=r, column=1, value=k).font = SECTION_FONT
        ws.cell(row=r, column=1).fill = SECTION_FILL
        ws.cell(row=r, column=1).border = BORDER
        ws.cell(row=r, column=1).alignment = LEFT
        ws.cell(row=r, column=2, value=v).font = BODY_FONT
        ws.cell(row=r, column=2).alignment = LEFT
        ws.cell(row=r, column=2).border = BORDER

    ws.cell(row=start + len(meta) + 2, column=1, value="■ シート構成").font = SECTION_FONT
    sheets_desc = [
        ("1. 機能一覧表", "システムが提供する機能の一覧（基本設計）"),
        ("2. API仕様書", "CLIインターフェースと内部モジュール関数の詳細仕様"),
        ("3. テーブル定義書", "JSON設定ファイルのスキーマ定義（DB相当）"),
        ("4. エラー・ログ定義書", "エラーメッセージ、エラーコード、出力タイミング"),
        ("5. アーキテクチャ図解", "Mermaid記法による各種図（構成図・ER図・シーケンス図）"),
    ]
    tbl_start = start + len(meta) + 3
    write_table(ws, tbl_start, ["シート名", "内容"], sheets_desc)


# ============================================================
# シート 2: 機能一覧表
# ============================================================
def build_feature_list(wb: Workbook) -> None:
    ws = wb.create_sheet("機能一覧表")
    set_column_widths(ws, [10, 28, 55, 20, 22, 20])

    ws.cell(row=1, column=1, value="■ 機能一覧表（基本設計）").font = TITLE_FONT
    ws.cell(row=2, column=1, value="本ツールが提供する機能を実装モジュール単位で列挙する。").font = BODY_FONT

    headers = ["機能ID", "機能名", "機能概要", "対象ユーザー", "実装モジュール", "実装関数"]
    rows = [
        ["F-001", "CLI引数解析",
         "--input / --site / --year / --output / --list-sites を受け付け、引数不備時にはargparseの標準エラー出力でユーザに通知する",
         "CLI利用者", "main.py", "build_parser() / main()"],
        ["F-002", "サイト一覧表示",
         "config/site_profiles/ 配下の*.jsonを走査し、_で始まるファイルを除外した上でsite_keyとdisplay_nameをコンソール表示する",
         "CLI利用者", "main.py", "list_available_sites()"],
        ["F-003", "サイトプロファイル読込",
         "指定されたサイト種別（例: rakuten）に対応するJSONファイルをロードし、カラムマッピング/エンコーディング/デフォルト販売元を取得する",
         "システム内部", "core/loader.py", "load_site_profile()"],
        ["F-004", "医薬品辞書読込",
         "config/medicine_dict/brands.json を読み込み、対象ブランドリストと除外キーワードリストを取得する",
         "システム内部", "core/matcher.py", "load_medicine_dict()"],
        ["F-005", "CSV読込・正規化",
         "サイト別カラムマッピングを適用してCSVを読み込み、内部統一カラム（order_date, product_name, unit_price, quantity, seller）に変換する。カンマ付き金額、全角￥、不正行を自動処理。",
         "システム内部", "core/loader.py", "load_csv()"],
        ["F-006", "年度フィルタ",
         "--year が指定された場合、注文日のうち該当年のレコードのみを残す",
         "CLI利用者", "core/loader.py", "filter_by_year()"],
        ["F-007", "商品名正規化",
         "NFKC正規化（全角英数→半角）、スペース統一、小文字化により商品名の表記ゆれを吸収する",
         "システム内部", "core/matcher.py", "normalize_text()"],
        ["F-008", "単品判定",
         "ブランド名との部分一致 → 除外キーワードチェックの順で3段階判定を返す（対象/要確認/対象外）",
         "システム内部", "core/matcher.py", "judge_product()"],
        ["F-009", "DataFrame一括判定",
         "全レコードに単品判定を適用し、判定列を付与。「対象外」を除外した結果を返す",
         "システム内部", "core/matcher.py", "apply_judgement()"],
        ["F-010", "xlsx出力",
         "判定済みDataFrameを書式付きxlsxに出力。ヘッダ装飾/要確認行の黄色ハイライト/合計行/列幅/数値フォーマットを自動適用",
         "CLI利用者", "core/exporter.py", "export_xlsx()"],
        ["F-011", "実行サマリ表示",
         "読込件数、年度フィルタ後件数、対象/要確認件数、合計金額、要確認件数の警告をコンソールに出力",
         "CLI利用者", "main.py", "main()"],
        ["F-012", "PyInstaller対応",
         "sys.frozen を判定してPROJECT_ROOTをsys.executableの親に切り替え、EXE配布時にconfig/を外部フォルダから読めるようにする",
         "配布利用者", "main.py", "（モジュール冒頭）"],
    ]
    write_table(ws, 4, headers, rows)


# ============================================================
# シート 3: API仕様書
# ============================================================
def build_api_spec(wb: Workbook) -> None:
    ws = wb.create_sheet("API仕様書")
    set_column_widths(ws, [8, 28, 20, 45, 35, 35])

    ws.cell(row=1, column=1, value="■ API仕様書（詳細設計）").font = TITLE_FONT
    ws.cell(row=2, column=1,
            value="本ツールはREST APIを持たないCLIツールのため、(1)CLIインターフェース と (2)内部モジュール関数API を記載する。").font = BODY_FONT

    # --- CLIインターフェース ---
    row = 4
    ws.cell(row=row, column=1, value="【1】CLIインターフェース").font = SECTION_FONT
    ws.cell(row=row, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1

    cli_headers = ["ID", "オプション", "短縮形", "説明", "型 / デフォルト", "バリデーション"]
    cli_rows = [
        ["C-01", "--input", "-i", "入力CSVファイルのパス", "Path / 必須（--list-sites 時を除く）", "ファイル存在チェック（main.py L102）"],
        ["C-02", "--site", "-s", "ECサイト種別（サイトキー）", "str / デフォルト 'amazon'", "config/site_profiles/<キー>.json の存在チェック（main.py L107）"],
        ["C-03", "--year", "-y", "対象年でフィルタ（例: 2025）", "int / デフォルト None（フィルタ無）", "（型のみargparseで検証）"],
        ["C-04", "--output", "-o", "出力xlsxファイルのパス", "Path / デフォルト 'selfmed_result.xlsx'", "親ディレクトリの存在は呼出側責任"],
        ["C-05", "--list-sites", "（なし）", "対応サイト一覧を表示して終了", "bool / デフォルト False", "trueの場合 --input は不要"],
        ["C-06", "--help / -h", "-h", "ヘルプ表示（argparse標準）", "（argparse自動）", "—"],
    ]
    row = write_table(ws, row, cli_headers, cli_rows)
    row += 2

    # --- CLIコマンド実行パターン ---
    ws.cell(row=row, column=1, value="【2】代表的なCLI実行パターン").font = SECTION_FONT
    ws.cell(row=row, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1
    pattern_headers = ["ID", "ユースケース", "コマンド例", "期待挙動", "exit code", "備考"]
    pattern_rows = [
        ["P-01", "対応サイト一覧表示", "main.py --list-sites",
         "サイトキーとdisplay_nameを一覧表示して正常終了", "0", "設定確認用"],
        ["P-02", "Amazon CSV を全年度処理", "main.py -i amazon.csv",
         "amazon.jsonでCSVを正規化→判定→selfmed_result.xlsx 出力", "0", "デフォルト動作"],
        ["P-03", "楽天 2025年のみ処理", "main.py -i rakuten.csv -s rakuten -y 2025",
         "rakuten.jsonで読込→2025年フィルタ→判定→xlsx出力", "0", "年度フィルタ"],
        ["P-04", "入力ファイル未存在", "main.py -i missing.csv",
         "stderrにエラー出力して異常終了", "1", "main.py L103-104"],
        ["P-05", "未定義サイト指定", "main.py -i x.csv -s unknown",
         "stderrにエラー出力して異常終了", "1", "main.py L108-111"],
        ["P-06", "必須カラム不足CSV", "main.py -i malformed.csv",
         "loader.load_csv() が ValueError を送出（main.py 未捕捉）", "非0", "F-005参照"],
    ]
    row = write_table(ws, row, pattern_headers, pattern_rows)
    row += 2

    # --- 内部モジュール関数API ---
    ws.cell(row=row, column=1, value="【3】内部モジュール関数API").font = SECTION_FONT
    ws.cell(row=row, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1

    mod_headers = ["ID", "モジュール.関数", "シグネチャ", "処理概要", "入力", "出力"]
    mod_rows = [
        ["M-01", "core.loader.load_site_profile",
         "load_site_profile(profile_path: Path) -> dict",
         "サイトプロファイルJSONを読み込む",
         "profile_path: JSONファイルパス",
         "dict（columns, encoding, date_format等）"],
        ["M-02", "core.loader.load_csv",
         "load_csv(csv_path: Path, profile: dict) -> pd.DataFrame",
         "CSVを読み込み、サイトプロファイルに基づき統一カラムに変換。カンマ付き金額/全角￥を除去し、不正行を除外",
         "csv_path: CSVパス / profile: サイトプロファイル辞書",
         "統一カラム(order_date, product_name, unit_price, quantity, seller)のDataFrame"],
        ["M-03", "core.loader.filter_by_year",
         "filter_by_year(df: pd.DataFrame, year: int) -> pd.DataFrame",
         "指定年の注文のみを抽出",
         "df: 統一カラムDataFrame / year: 対象年",
         "フィルタ済みDataFrame"],
        ["M-04", "core.matcher.load_medicine_dict",
         "load_medicine_dict(dict_path: Path) -> dict",
         "医薬品辞書JSONを読み込む",
         "dict_path: JSONファイルパス",
         "dict（brands, exclude_keywords）"],
        ["M-05", "core.matcher.normalize_text",
         "normalize_text(text: str) -> str",
         "NFKC正規化 + スペース統一 + 前後空白除去 + 小文字化",
         "text: 元テキスト",
         "正規化済みテキスト"],
        ["M-06", "core.matcher.judge_product",
         "judge_product(product_name: str, brands: list[str], exclude_keywords: list[str]) -> str",
         "商品名を3段階判定。ブランド部分一致→除外キーワードチェック",
         "商品名 / ブランドリスト / 除外リスト",
         "'対象' / '要確認' / '対象外' のいずれか"],
        ["M-07", "core.matcher.apply_judgement",
         "apply_judgement(df: pd.DataFrame, medicine_dict: dict) -> pd.DataFrame",
         "全レコードに判定を付与し、'対象外'を除外",
         "df: 統一DataFrame / medicine_dict: 辞書",
         "判定列付きDataFrame（対象外除外済）"],
        ["M-08", "core.exporter.export_xlsx",
         "export_xlsx(df: pd.DataFrame, output_path: Path) -> Path",
         "判定済みDataFrameをxlsxに出力（ヘッダ装飾/要確認ハイライト/合計行/列幅）",
         "df: 判定済みDataFrame / output_path: 出力パス",
         "出力されたファイルパス"],
        ["M-09", "main.list_available_sites",
         "list_available_sites() -> list[tuple[str, str]]",
         "site_profilesディレクトリを走査し、_で始まるファイルを除外して(site_key, display_name)のリストを返す",
         "（引数なし）",
         "[(key, name), ...]"],
        ["M-10", "main.build_parser",
         "build_parser() -> argparse.ArgumentParser",
         "argparseの引数定義を構築して返す",
         "（引数なし）",
         "ArgumentParserインスタンス"],
        ["M-11", "main.main",
         "main(args: list[str] | None = None) -> int",
         "エントリポイント。引数解析→読込→フィルタ→判定→出力の全パイプラインを実行",
         "args: コマンドライン引数（Noneでsys.argvを使用）",
         "終了コード（0=正常 / 1=エラー）"],
    ]
    write_table(ws, row, mod_headers, mod_rows)


# ============================================================
# シート 4: テーブル定義書（JSON設定ファイル）
# ============================================================
def build_table_def(wb: Workbook) -> None:
    ws = wb.create_sheet("テーブル定義書")
    set_column_widths(ws, [8, 20, 15, 12, 15, 45])

    ws.cell(row=1, column=1, value="■ テーブル定義書（詳細設計）").font = TITLE_FONT
    ws.cell(row=2, column=1,
            value="本ツールはRDBを持たない。データ定義はJSON設定ファイルが担うため、各JSONのスキーマを「テーブル定義」として記載する。").font = BODY_FONT

    row = 4

    # --- サイトプロファイル ---
    ws.cell(row=row, column=1,
            value="【T1】site_profile （config/site_profiles/<site_key>.json）").font = SECTION_FONT
    ws.cell(row=row, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1
    ws.cell(row=row, column=1,
            value="論理名: ECサイトごとのCSVカラムマッピング定義 / 物理名: <site_key>.json / 主キー: ファイル名（site_key）").font = BODY_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1

    hdr = ["No", "カラム名(キー)", "論理名", "型", "制約", "説明"]
    rows = [
        [1, "_comment", "コメント", "string", "任意", "人間向けの備考。プログラムは無視する"],
        [2, "display_name", "表示名", "string", "任意（推奨）", "--list-sites で表示される日本語名。省略時は site_key で代替"],
        [3, "encoding", "文字エンコーディング", "string", "任意（デフォルト utf-8）", "CSV読込時のエンコーディング指定（utf-8 / shift_jis / cp932 等）"],
        [4, "columns", "カラムマッピング", "object", "必須", "CSV上のヘッダー名から内部統一カラムへの変換辞書（下記サブフィールド）"],
        [5, "columns.order_date", "注文日カラム名", "string", "必須", "実CSV上の注文日カラムの見出し名（例: '注文日'）"],
        [6, "columns.product_name", "商品名カラム名", "string", "必須", "実CSV上の商品名カラムの見出し名"],
        [7, "columns.unit_price", "金額カラム名", "string", "必須", "実CSV上の金額カラムの見出し名"],
        [8, "columns.quantity", "数量カラム名", "string", "任意", "実CSV上の数量カラム名。無ければ1として扱う（loader.py L64-65）"],
        [9, "columns.seller", "販売元カラム名", "string", "任意", "実CSV上の販売元カラム名。無ければ default_seller で補完"],
        [10, "default_seller", "デフォルト販売元", "string", "任意（デフォルト '不明'）", "seller カラムが無い／欠損時に出力される支払先名"],
        [11, "date_format", "日付フォーマット", "string", "任意（デフォルト %Y-%m-%d）", "pandasのto_datetime用フォーマット文字列"],
    ]
    row = write_table(ws, row, hdr, rows)
    row += 2

    # --- 医薬品辞書 ---
    ws.cell(row=row, column=1,
            value="【T2】medicine_dict （config/medicine_dict/brands.json）").font = SECTION_FONT
    ws.cell(row=row, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1
    ws.cell(row=row, column=1,
            value="論理名: セルフメディケーション税制対象品ブランド辞書 / 物理名: brands.json / 主キー: （なし・単一ファイル）").font = BODY_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1
    rows = [
        [1, "brands", "対象ブランド一覧", "array[string]", "必須", "対象医薬品のブランド名配列。部分一致で商品名と照合される"],
        [2, "exclude_keywords", "除外キーワード一覧", "array[string]", "任意（デフォルト []）", "この単語を含む商品は判定を '要確認' に降格（例: 'ルナ','キッズ'）"],
    ]
    row = write_table(ws, row, hdr, rows)
    row += 2

    # --- 内部統一DataFrame ---
    ws.cell(row=row, column=1,
            value="【T3】unified_dataframe （loader→matcher→exporter 間で受け渡すメモリ上データ構造）").font = SECTION_FONT
    ws.cell(row=row, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1
    ws.cell(row=row, column=1,
            value="論理名: 正規化済み購入履歴 / 定義場所: core/loader.py:UNIFIED_COLUMNS").font = BODY_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1
    rows = [
        [1, "order_date", "注文日", "datetime64", "NOT NULL", "loader で to_datetime 変換。NaTの行は除外される"],
        [2, "product_name", "商品名", "string", "NOT NULL", "元CSVの商品名をそのまま保持（正規化は判定時のみ適用）"],
        [3, "unit_price", "支払金額", "float→int", "NOT NULL", "カンマ・￥を除去してから数値化。出力時に int 化"],
        [4, "quantity", "数量", "int", "NOT NULL（デフォルト1）", "loader で int 型に変換。エクスポート対象カラムではない"],
        [5, "seller", "販売元", "string", "NOT NULL", "欠損時は profile.default_seller で補完"],
        [6, "判定", "判定結果", "string", "matcher 適用後に追加", "値域: '対象' / '要確認'（'対象外' は行ごと除外）"],
    ]
    row = write_table(ws, row, hdr, rows)
    row += 2

    # --- 出力xlsx ---
    ws.cell(row=row, column=1,
            value="【T4】output_xlsx （core/exporter.py:OUTPUT_COLUMNS）").font = SECTION_FONT
    ws.cell(row=row, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1
    ws.cell(row=row, column=1,
            value="論理名: 確定申告用明細出力 / シート名: 'セルフメディケーション明細'").font = BODY_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1
    rows = [
        [1, "A: 支払先の名称", "seller 由来", "string", "—", "列幅20。ヘッダ青背景/白文字"],
        [2, "B: 医薬品の名称", "product_name 由来", "string", "—", "列幅35"],
        [3, "C: 支払った金額", "unit_price 由来", "int", "数値書式 #,##0", "列幅15。要確認行は黄色(#FFF2CC)背景"],
        [4, "D: 購入日", "order_date 由来（%Y-%m-%d）", "string", "—", "列幅14"],
        [5, "E: 判定", "対象 / 要確認", "string", "—", "列幅10"],
        [6, "最終行: 合計", "unit_price の合計", "int", "合計ラベル(B列) + 金額(C列)", "Bold / 罫線付き"],
    ]
    write_table(ws, row, hdr, rows)


# ============================================================
# シート 5: エラー・ログ定義書
# ============================================================
def build_error_log(wb: Workbook) -> None:
    ws = wb.create_sheet("エラー・ログ定義")
    set_column_widths(ws, [10, 28, 10, 55, 28, 30])

    ws.cell(row=1, column=1, value="■ エラー・ログ定義書（詳細設計）").font = TITLE_FONT
    ws.cell(row=2, column=1,
            value="本ツールでは logging モジュールを使用せず、print/print(file=sys.stderr)/raise によってユーザへ通知する。下記は現在の実装から抽出したメッセージ一覧。").font = BODY_FONT

    row = 4
    ws.cell(row=row, column=1, value="【1】エラー／通知メッセージ一覧").font = SECTION_FONT
    ws.cell(row=row, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1

    hdr = ["エラーID", "レベル", "出力先", "メッセージ（原文）", "発生箇所", "発生タイミング / 対処"]
    rows = [
        ["E-001", "ERROR", "stderr",
         "エラー: 入力ファイルが見つかりません: <path>",
         "main.py L103",
         "--input で指定されたCSVが存在しない時に表示し exit 1。対処: パスを確認"],
        ["E-002", "ERROR", "stderr",
         "エラー: サイトプロファイルが見つかりません: <path>\n  対応サイト一覧は `python main.py --list-sites` で確認できます。",
         "main.py L109-110",
         "--site で指定したキーのJSONが存在しない時。対処: --list-sites で確認"],
        ["E-003", "EXCEPTION", "例外送出（ValueError）",
         "必須カラムが見つかりません: [...]。CSVのカラム: [...]。site_profileのマッピングを確認してください。",
         "core/loader.py L57-61",
         "CSVのヘッダに order_date/product_name/unit_price のいずれかが見つからない時。対処: site_profile.columns を編集"],
        ["E-004", "EXCEPTION", "例外送出（FileNotFoundError）",
         "[Errno 2] No such file or directory: <path>",
         "core/loader.py L19（open）",
         "site_profile.json が存在しない時（通常はmain.pyで前段チェック済み）"],
        ["E-005", "EXCEPTION", "例外送出（FileNotFoundError）",
         "[Errno 2] No such file or directory: <path>",
         "core/matcher.py L17（open）",
         "brands.json が存在しない時。対処: config/medicine_dict/brands.json を配置"],
        ["E-006", "EXCEPTION", "例外送出（UnicodeDecodeError / pd.errors.ParserError）",
         "（各ライブラリの標準メッセージ）",
         "core/loader.py L39（pd.read_csv）",
         "encoding 不一致やCSV構造不正時。対処: site_profile.encoding を確認"],
        ["E-007", "EXCEPTION", "例外送出（KeyError）",
         "'columns'",
         "core/loader.py L34",
         "site_profile JSON に columns キーが欠落している時。対処: JSONスキーマを確認（T1参照）"],
        ["E-008", "EXCEPTION", "例外送出（KeyError）",
         "'brands'",
         "core/matcher.py L78",
         "brands.json に brands キーが欠落している時。対処: JSONスキーマを確認（T2参照）"],
        ["E-009", "ERROR (argparse)", "stderr",
         "usage: main.py [...] main.py: error: --input は必須です（--list-sites 以外の場合）",
         "main.py L99",
         "--input と --list-sites の両方が未指定の時。対処: いずれかを指定"],
    ]
    row = write_table(ws, row, hdr, rows)
    row += 2

    # --- 正常系ログ ---
    ws.cell(row=row, column=1, value="【2】正常系のコンソール出力（print → stdout）").font = SECTION_FONT
    ws.cell(row=row, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1
    hdr2 = ["ログID", "レベル", "出力先", "メッセージ", "発生箇所", "出力タイミング"]
    log_rows = [
        ["L-001", "INFO", "stdout", "対応ECサイト一覧:\n  <key> <display_name> ...", "main.py L92-94", "--list-sites 実行時"],
        ["L-002", "INFO", "stdout", "読み込み中: <input_path>", "main.py L120", "CSV読込開始時"],
        ["L-003", "INFO", "stdout", "  → <N> 件の注文を読み込みました", "main.py L122", "CSV読込完了時"],
        ["L-004", "INFO", "stdout", "  → <year>年の注文: <N> 件", "main.py L127", "--year 指定時、年度フィルタ後"],
        ["L-005", "INFO", "stdout", "  → 対象: <N> 件 / 要確認: <N> 件", "main.py L133", "判定完了時"],
        ["L-006", "INFO", "stdout", "対象となる医薬品は見つかりませんでした。", "main.py L137", "判定後の件数が0の時（exit 0）"],
        ["L-007", "INFO", "stdout", "出力完了: <output_path>", "main.py L143", "xlsx書き出し完了時"],
        ["L-008", "INFO", "stdout", "合計金額: <amount>円", "main.py L144", "xlsx書き出し完了時"],
        ["L-009", "WARN", "stdout", "※ 「要確認」が <N> 件あります。出力ファイルの黄色行を確認してください。", "main.py L146", "要確認件数が1件以上の時"],
    ]
    row = write_table(ws, row, hdr2, log_rows)
    row += 2

    # --- 終了コード ---
    ws.cell(row=row, column=1, value="【3】プロセス終了コード").font = SECTION_FONT
    ws.cell(row=row, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1
    hdr3 = ["コード", "意味", "発生条件", "発生箇所", "備考", "—"]
    code_rows = [
        [0, "正常終了", "処理成功、または対象0件でもファイルなしとして正常終了", "main.py L95/138/148", "—", ""],
        [1, "入力エラー", "入力ファイル or サイトプロファイルが存在しない", "main.py L104/111", "ユーザの指定ミスを想定", ""],
        ["非0", "例外終了", "loader/matcher/exporter が例外を送出しmainで未捕捉", "—", "ValueError / FileNotFoundError 等", ""],
        [2, "argparse終了", "必須引数不足（--input と --list-sites の両方欠）", "argparse 標準動作", "parser.error() 経由", ""],
    ]
    write_table(ws, row, hdr3, code_rows)


# ============================================================
# シート 6: アーキテクチャ図解（Mermaid記法）
# ============================================================
def build_mermaid(wb: Workbook) -> None:
    ws = wb.create_sheet("アーキテクチャ図解")
    set_column_widths(ws, [25, 90])

    ws.cell(row=1, column=1, value="■ アーキテクチャ図解（Mermaid記法）").font = TITLE_FONT
    ws.cell(row=2, column=1,
            value="各図は Mermaid 記法で記載。https://mermaid.live/ 等で貼り付けて描画可能。").font = BODY_FONT

    diagrams = [
        ("1. システム構成図 (Component Diagram)", """graph LR
    User[CLI利用者] -->|コマンド実行| CLI[main.py<br/>argparse]
    CLI --> Loader[core/loader.py<br/>CSV読込・正規化]
    CLI --> Matcher[core/matcher.py<br/>医薬品判定]
    CLI --> Exporter[core/exporter.py<br/>xlsx出力]

    Loader -->|読込| SiteProfile[(config/site_profiles/<br/>*.json)]
    Loader -->|読込| InputCSV[(入力CSV)]
    Matcher -->|読込| MedDict[(config/medicine_dict/<br/>brands.json)]
    Exporter -->|書込| OutputXlsx[(出力xlsx)]

    Loader -.DataFrame.-> Matcher
    Matcher -.判定付DF.-> Exporter"""),

        ("2. データフロー図 (Data Flow)", """flowchart TD
    Start([開始]) --> ArgParse[引数解析<br/>build_parser]
    ArgParse --> ListCheck{--list-sites?}
    ListCheck -->|Yes| ListSites[list_available_sites]
    ListSites --> End([終了 exit 0])
    ListCheck -->|No| InputCheck{入力ファイル<br/>存在?}
    InputCheck -->|No| ErrExit1([stderr出力<br/>exit 1])
    InputCheck -->|Yes| ProfileCheck{プロファイル<br/>存在?}
    ProfileCheck -->|No| ErrExit2([stderr出力<br/>exit 1])
    ProfileCheck -->|Yes| LoadProfile[load_site_profile]
    LoadProfile --> LoadDict[load_medicine_dict]
    LoadDict --> LoadCSV[load_csv<br/>正規化]
    LoadCSV --> YearCheck{--year 指定?}
    YearCheck -->|Yes| Filter[filter_by_year]
    YearCheck -->|No| Judge
    Filter --> Judge[apply_judgement<br/>3段階判定]
    Judge --> ResultCheck{対象件数 > 0?}
    ResultCheck -->|No| NoResult[メッセージ表示] --> End
    ResultCheck -->|Yes| Export[export_xlsx]
    Export --> Summary[サマリ表示<br/>合計/要確認警告]
    Summary --> End"""),

        ("3. ER図 (設定ファイル構造)", """erDiagram
    SITE_PROFILE ||--|| COLUMNS_MAP : contains
    SITE_PROFILE {
        string site_key PK "ファイル名"
        string display_name
        string encoding
        string default_seller
        string date_format
    }
    COLUMNS_MAP {
        string order_date "CSVヘッダ名"
        string product_name "CSVヘッダ名"
        string unit_price "CSVヘッダ名"
        string quantity "CSVヘッダ名(任意)"
        string seller "CSVヘッダ名(任意)"
    }

    MEDICINE_DICT ||--o{ BRAND : has
    MEDICINE_DICT ||--o{ EXCLUDE_KEYWORD : has
    MEDICINE_DICT {
        string file "brands.json"
    }
    BRAND {
        string name "対象ブランド名"
    }
    EXCLUDE_KEYWORD {
        string keyword "除外語"
    }

    UNIFIED_DF }o--|| SITE_PROFILE : generated_from
    UNIFIED_DF }o--|| MEDICINE_DICT : judged_by
    UNIFIED_DF {
        datetime order_date
        string product_name
        int unit_price
        int quantity
        string seller
        string judgement "対象/要確認"
    }"""),

        ("4. シーケンス図 (E2E処理フロー)", """sequenceDiagram
    actor User as CLI利用者
    participant Main as main.py
    participant Loader as core/loader
    participant Matcher as core/matcher
    participant Exporter as core/exporter
    participant FS as ファイルシステム

    User->>Main: main.py -i x.csv -s rakuten -y 2025
    Main->>Main: build_parser / parse_args
    Main->>FS: input.exists()
    FS-->>Main: True
    Main->>FS: profile_path.exists()
    FS-->>Main: True
    Main->>Loader: load_site_profile(path)
    Loader->>FS: read rakuten.json
    FS-->>Loader: profile dict
    Loader-->>Main: profile
    Main->>Matcher: load_medicine_dict(path)
    Matcher->>FS: read brands.json
    FS-->>Matcher: dict
    Matcher-->>Main: medicine_dict
    Main->>Loader: load_csv(path, profile)
    Loader->>FS: read CSV
    FS-->>Loader: raw DataFrame
    Loader->>Loader: カラムマッピング/型変換/クレンジング
    Loader-->>Main: unified DataFrame
    Main->>Loader: filter_by_year(df, 2025)
    Loader-->>Main: filtered DataFrame
    Main->>Matcher: apply_judgement(df, dict)
    loop 各レコード
        Matcher->>Matcher: normalize_text
        Matcher->>Matcher: judge_product
    end
    Matcher-->>Main: judged DataFrame (対象外除外)
    Main->>Exporter: export_xlsx(df, output_path)
    Exporter->>FS: write xlsx (書式付き)
    FS-->>Exporter: saved
    Exporter-->>Main: output_path
    Main-->>User: サマリ表示 / exit 0"""),

        ("5. 判定ロジック図 (State / Decision)", """stateDiagram-v2
    [*] --> 入力商品名
    入力商品名 --> 正規化: normalize_text<br/>NFKC/スペース/小文字化
    正規化 --> ブランド照合
    ブランド照合 --> 対象外: brands いずれもマッチせず
    ブランド照合 --> 除外キーワード判定: 部分一致あり
    除外キーワード判定 --> 要確認: exclude_keywords 含む
    除外キーワード判定 --> 対象: 含まない
    対象外 --> [*]: 出力から除外
    対象 --> [*]: xlsx に出力
    要確認 --> [*]: xlsx に黄色ハイライトで出力"""),

        ("6. ディレクトリ構成図", """graph TD
    Root[selfmed-tax-tool/]
    Root --> Main[main.py]
    Root --> Req[requirements.txt]
    Root --> Readme[README.md]
    Root --> Spec[build_exe.spec]
    Root --> Config[config/]
    Root --> Core[core/]
    Root --> Tests[tests/]
    Root --> Dist[dist/]

    Config --> SP[site_profiles/]
    Config --> MD[medicine_dict/]
    SP --> AMZ[amazon.json]
    SP --> RKT[rakuten.json]
    SP --> OTH[... 他11サイト<br/>+ _template.json]
    MD --> BRN[brands.json]

    Core --> LD[loader.py]
    Core --> MT[matcher.py]
    Core --> EX[exporter.py]
    Core --> INIT[__init__.py]

    Tests --> TL[test_loader.py]
    Tests --> TM[test_matcher.py]
    Tests --> TE[test_exporter.py]
    Tests --> FX[fixtures/]

    Dist --> EXE[SelfmedTaxTool.exe]
    Dist --> DIST_CFG[config/]
    Dist --> DOC[SelfmedTaxTool説明書.txt]"""),
    ]

    row = 4
    for title, code in diagrams:
        ws.cell(row=row, column=1, value=title).font = SECTION_FONT
        ws.cell(row=row, column=1).fill = SECTION_FILL
        ws.cell(row=row, column=1).border = BORDER
        ws.cell(row=row, column=1).alignment = LEFT
        ws.cell(row=row, column=2, value=code).font = Font(name="Consolas", size=9)
        ws.cell(row=row, column=2).alignment = LEFT
        ws.cell(row=row, column=2).border = BORDER
        line_count = code.count("\n") + 1
        ws.row_dimensions[row].height = max(20, line_count * 13)
        row += 2


def main():
    wb = Workbook()
    build_cover(wb)
    build_feature_list(wb)
    build_api_spec(wb)
    build_table_def(wb)
    build_error_log(wb)
    build_mermaid(wb)
    output = "design_document.xlsx"
    wb.save(output)
    print(f"生成完了: {output}")
    print(f"シート数: {len(wb.sheetnames)}")
    for name in wb.sheetnames:
        print(f"  - {name}")


if __name__ == "__main__":
    main()
