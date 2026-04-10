"""xlsx出力モジュール。

判定済みDataFrameをセルフメディケーション税制の明細書に
転記しやすいExcelファイルとして出力する。
"""

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# 出力カラムの定義（内部名 → 表示名）
OUTPUT_COLUMNS = {
    "seller": "支払先の名称",
    "product_name": "医薬品の名称",
    "unit_price": "支払った金額",
    "order_date": "購入日",
    "判定": "判定",
}


def export_xlsx(df: pd.DataFrame, output_path: Path) -> Path:
    """判定済みDataFrameをxlsxファイルとして出力する。

    Args:
        df: 判定列が付与されたDataFrame。
        output_path: 出力先ファイルパス。

    Returns:
        出力ファイルのパス。
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "セルフメディケーション明細"

    # --- スタイル定義 ---
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    review_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    total_font = Font(bold=True, size=11)

    # --- ヘッダー行 ---
    headers = list(OUTPUT_COLUMNS.values())
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # --- データ行 ---
    internal_cols = list(OUTPUT_COLUMNS.keys())
    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        for col_idx, col_name in enumerate(internal_cols, 1):
            value = row[col_name]
            # 日付をフォーマット
            if col_name == "order_date" and pd.notna(value):
                value = value.strftime("%Y-%m-%d")
            # 金額を整数化
            if col_name == "unit_price" and pd.notna(value):
                value = int(value)

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

            # 「要確認」行は背景色でハイライト
            if row.get("判定") == "要確認":
                cell.fill = review_fill

    # --- 合計行 ---
    data_end_row = len(df) + 1
    total_row = data_end_row + 1
    price_col = internal_cols.index("unit_price") + 1

    # 「合計」ラベル
    label_cell = ws.cell(row=total_row, column=price_col - 1, value="合計")
    label_cell.font = total_font
    label_cell.alignment = Alignment(horizontal="right")
    label_cell.border = thin_border

    # 合計金額
    total_amount = int(df["unit_price"].sum()) if len(df) > 0 else 0
    total_cell = ws.cell(row=total_row, column=price_col, value=total_amount)
    total_cell.font = total_font
    total_cell.border = thin_border

    # --- 列幅の自動調整 ---
    column_widths = {
        "支払先の名称": 20,
        "医薬品の名称": 35,
        "支払った金額": 15,
        "購入日": 14,
        "判定": 10,
    }
    for col_idx, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = column_widths.get(header, 15)

    # --- 金額列の表示形式 ---
    for row_idx in range(2, total_row + 1):
        ws.cell(row=row_idx, column=price_col).number_format = '#,##0'

    wb.save(output_path)
    return output_path
